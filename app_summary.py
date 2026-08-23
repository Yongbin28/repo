import os
import shutil
import tempfile
import pandas as pd
from pathlib import Path
import warnings
import re
import sys # Added for sys.path and sys.executable

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Paths
# --- The current directory is added to the system path to allow importing sibling modules. ---
if getattr(sys, 'frozen', False):
    CURRENT_DIR = Path(sys.executable).parent.resolve()
else:
    CURRENT_DIR = Path(__file__).parent.resolve()

if str(CURRENT_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_DIR))

DATASET_ROOT = CURRENT_DIR / "dataset"

OUTPUT_FILE = CURRENT_DIR / "Pipeline_Status_Summary.xlsx"
TESTER_MAPPING_FILE = DATASET_ROOT / "Unique_Generics_Testers.csv"

def get_tester_mapping():
    mapping = {}
    if TESTER_MAPPING_FILE.exists():
        try:
            df = pd.read_csv(TESTER_MAPPING_FILE)
            for _, row in df.iterrows():
                gen = str(row.get('Generic', '')).strip()
                tstr = str(row.get('Tester', '')).strip()
                if gen and tstr:
                    mapping[gen] = tstr
        except Exception as e:
            print(f"Error reading mapping: {e}")
    return mapping

def read_data_safely(file_path):
    """Attempts to read a CSV or XLSX safely. Skips unhydrated OneDrive placeholders."""
    try:
        if file_path.stat().st_size == 0:
            print(f"  -> Skipping {file_path.name} (0 bytes / placeholder)")
            return None, None
            
        if file_path.suffix.lower() == '.csv':
            return pd.read_csv(file_path), None
        elif file_path.suffix.lower() == '.xlsx':
            try:
                # Read all of Summary sheet
                df_all = pd.read_excel(file_path, sheet_name='Summary', header=None)
                
                val_rows = pd.NA
                # Find validation rows in the top metadata
                for _, row in df_all.head(15).iterrows():
                    val = str(row.values[0])
                    if 'Val rows:' in val:
                        m = re.search(r'Val rows:\s*(\d+)', val)
                        if m:
                            val_rows = int(m.group(1))
                        break
                
                # Find the row that contains 'model' header
                header_row = -1
                for i, row in df_all.iterrows():
                    if 'model' in row.values and 'r2' in row.values:
                        header_row = i
                        break
                
                if header_row != -1:
                    df = pd.read_excel(file_path, sheet_name='Summary', header=header_row)
                    return df, val_rows
                return None, None
            except Exception as e:
                print(f"  -> Extracting from {file_path.name} failed: {e}")
                return None, None
                
    except OSError as e:
        if e.errno == 22: # Invalid Argument (OneDrive Cloud Placeholder)
            print(f"  -> Skipping {file_path.name}: OneDrive placeholder not synced locally (Errno 22).")
            return None, None
        raise e
    except Exception as e:
        print(f"  -> Error reading {file_path.name}: {e}")
        return None, None

def extract_model_accuracies(generics_list=None, generic_logs=None):
    results = []
    validation_sheets = {} # Dictionary to hold validation dataframes by generic
    
    if generics_list is None:
        generics_list = []
    if generic_logs is None:
        generic_logs = {}
        
    print(f"Scanning dataset for analytics in: {DATASET_ROOT}")
    
    # Ensure output directory exists
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    
    tester_mapping = get_tester_mapping()
    processed_generics = set()
    devices_processed = 0
    
    # If it is runStandalone from terminal without a generics list, scan the directory.
    if not generics_list:
        print("No generics_list provided (Running Standalone). Scanning dataset for processed devices...")
        for root, dirs, files in os.walk(DATASET_ROOT):
            root_path = Path(root)
            if root_path.name == "Model":
                # Add the generic folder name (parent of Model)
                generics_list.append(root_path.parent.name)
        
        # Remove duplicates from generic suffixes, keep base names
        scanned_generics = set()
        for g in generics_list:
            if '_' in g:
                scanned_generics.add(g.split('_')[0])
            else:
                scanned_generics.add(g)
        generics_list = sorted(list(scanned_generics))
        print(f"Discovered {len(generics_list)} generics to process.")

    for generic in generics_list:
        logs_for_generic = generic_logs.get(generic, [])
        log_str = "\n".join(logs_for_generic) if logs_for_generic else "None"
        
        # Look for the model folder
        tester_name = tester_mapping.get(generic, "Unknown")
        found_model = False
        
        # Need to search the dataset for the generic folder
        # Path structure: dataset / Family / Generic / Model / model_results...
        model_files = []
        val_df = None
        
        for root, dirs, files in os.walk(DATASET_ROOT):
            root_path = Path(root)
            # Find the Model folder for this generic
            if root_path.name == "Model" and (generic in root_path.parent.name):
                # Found the Model folder for this generic, collect results
                for f in files:
                    if (f.startswith("model_results") and f.endswith(".csv")) or (f.startswith("model_details") and f.endswith(".xlsx")):
                        model_files.append((root, f))
                        
                        # Extract the Validation sheet from the model_details xlsx
                        if f.startswith("model_details") and f.endswith(".xlsx"):
                            try:
                                xlsx_path = root_path / f
                                val_df = pd.read_excel(xlsx_path, sheet_name='Validation')
                            except Exception as e:
                                print(f"Could not load Validation sheet for {generic}: {e}")
                                
        # Sort so .xlsx is processed BEFORE .csv to prioritize getting Val Rows
        model_files.sort(key=lambda x: x[1].endswith('.csv'))
        
        for root, f in model_files:
            file_path = Path(root) / f
            try:
                parts = file_path.relative_to(DATASET_ROOT).parts
                if len(parts) >= 2:
                    tester_family = parts[0]
                    generic_folder = parts[1]
                    
                    # Update tester_name if mapping was unknown but path has info
                    if tester_name == "Unknown" and tester_family:
                        tester_name = tester_family

                    df, val_rows = read_data_safely(file_path)
                    if df is not None and not df.empty and 'r2' in df.columns:
                        top_models = df.head(3)
                        top_3_names = top_models['model'].tolist()
                        
                        for i, (_, row) in enumerate(top_models.iterrows()):
                            r2_val = row.get("r2", pd.NA)
                            
                            is_good = False
                            if pd.notna(r2_val):
                                try:
                                    is_good = float(r2_val) > 0.5
                                except ValueError:
                                    pass
                            
                            results.append({
                                "Tester Name": tester_name,
                                "Generic": generic,
                                "Partname": generic_folder,
                                "Rank": i + 1,
                                "Best Model": row.get("model", "Unknown"),
                                "Validation Rows": val_rows if pd.notna(val_rows) else "Unknown",
                                "R2": r2_val,
                                "RMSE": row.get("rmse", pd.NA),
                                "MAE": row.get("mae", pd.NA),
                                "R2 > 0.5": "Yes" if is_good else "No",
                                "Warnings & Errors": log_str,
                                "Details File": str(file_path.relative_to(DATASET_ROOT))
                            })
                            
                        processed_generics.add(generic)
                        if val_df is not None:
                            # Use a tuple key for uniqueness
                            validation_sheets[(generic, generic_folder)] = (val_df, top_3_names)
                        
                        devices_processed += 1
                        val_str = f"| ValRows: {val_rows}" if pd.notna(val_rows) else ""
                        print(f"Extracted device {devices_processed}: {generic} ({generic_folder}) [{tester_name}] ({len(top_models)} models) {val_str}")
                        found_model = True
                        break # Only need one successful read per generic
                        
            except Exception as e:
                print(f"Error processing path {file_path}: {e}")
                
        if not found_model:
            # Add an entry for the generic even if no model was created (Acts as an audit trail)
            results.append({
                "Tester Name": tester_name,
                "Generic": generic,
                "Partname": generic,
                "Rank": "N/A",
                "Best Model": "N/A",
                "Validation Rows": "N/A",
                "R2": pd.NA,
                "RMSE": pd.NA,
                "MAE": pd.NA,
                "R2 > 0.5": "N/A",
                "Warnings & Errors": log_str,
                "Details File": "N/A"
            })
            processed_generics.add(generic)
            devices_processed += 1
            print(f"Extracted device {devices_processed}: {generic} - Recorded as Processed (No Model)")
                    
    if results:
        summary_df = pd.DataFrame(results)
        
        # Check if the output file already exists to append/update models
        if OUTPUT_FILE.exists():
            try:
                print(f"\nMerging with existing `{OUTPUT_FILE.name}`...")
                existing_excel_file = pd.ExcelFile(OUTPUT_FILE)
                if 'Status' in existing_excel_file.sheet_names:
                    existing_df = existing_excel_file.parse('Status')
                elif 'Model Accuracies' in existing_excel_file.sheet_names:
                    # Support legacy sheet name migration
                    existing_df = existing_excel_file.parse('Model Accuracies')
                else:
                    existing_df = pd.DataFrame()
                
                # Need to drop older records for generics + testers + partnames that are being updated now
                keys = ["Tester Name", "Generic", "Partname"]
                
                # Ensure the columns exist in the existing df
                if all(k in existing_df.columns for k in keys):
                    # Find which generics/partname combos are in the NEW results, and remove them from the OLD results
                    updated_combinations = summary_df[keys].drop_duplicates()
                    
                    # Merge logic: keep rows in existing_df that do NOT match the ['Tester Name', 'Generic'] pairs in new summary_df
                    merged_keys = pd.merge(existing_df, updated_combinations, on=keys, how='left', indicator=True)
                    existing_filtered_df = existing_df[merged_keys['_merge'] == 'left_only']
                    
                    # Concat remaining old records with new records
                    summary_df = pd.concat([existing_filtered_df, summary_df], ignore_index=True)
                else:
                    # If columns don't match for some reason, just concat (append)
                    summary_df = pd.concat([existing_df, summary_df], ignore_index=True)
                    
            except Exception as e:
                print(f"Failed to merge existing Excel file (it may be corrupted or locked): {e}")
                print("Will overwrite with newly extracted data instead.")

        # Sort by Tester Name, Generic, and Rank (to keep top 3 models ordered)
        summary_df.sort_values(by=["Tester Name", "Generic", "Rank"], inplace=True)
        summary_df.reset_index(drop=True, inplace=True)
        
        total_devices_tracked = len(summary_df["Generic"].unique())
        
        # Calculate exactly how many devices actually produced models
        models_df = summary_df[summary_df["Best Model"] != "N/A"]
        
        # Only count rank 1 models when assessing "Total Models Generated" globally 
        # To match previous single-model logic 
        top1_models_df = models_df[models_df["Rank"] == 1]
        
        total_models = len(top1_models_df)
        good_models = (top1_models_df["R2 > 0.5"] == "Yes").sum()
        percentage = (good_models / total_models) * 100 if total_models > 0 else 0
        
        print(f"\n--- Summary ---")
        print(f"Total Devices Tracked: {total_devices_tracked}")
        print(f"Total Successful Models: {total_models}")
        print(f"Models with R2 > 0.5: {good_models} ({percentage:.1f}%)")
        print(f"---------------")
        
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Status', index=False)
            
            summary_stats = pd.DataFrame([{
                "Metric": "Total Devices Processed",
                "Value": total_devices_tracked
            }, {
                "Metric": "Total Models Generated",
                "Value": total_models
            }, {
                "Metric": "Models with R2 > 0.5",
                "Value": good_models
            }, {
                "Metric": "Percentage R2 > 0.5",
                "Value": f"{percentage:.1f}%"
            }])
            summary_stats.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format main sheets
            for sheet_name in ['Status', 'Summary']:
                if sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    df_ref = summary_df if sheet_name == 'Status' else summary_stats
                    
                    # Convert 'Status' into a formatted Excel Table with filters
                    if sheet_name == 'Status':
                        max_row = len(df_ref.index) + 1
                        max_col = len(df_ref.columns)
                        col_letter = get_column_letter(max_col)
                        ref_str = f"A1:{col_letter}{max_row}"
                        tab = Table(displayName="StatusTable", ref=ref_str)
                        # No styling/color for the main table (just standard plain text + filters)
                        style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                                            showLastColumn=False, showRowStripes=False, showColumnStripes=False)
                        tab.tableStyleInfo = style
                        worksheet.add_table(tab)
                    
                    # Setup column widths
                    for i, col in enumerate(df_ref.columns):
                        lengths = df_ref[col].dropna().astype(str).map(len)
                        val_len = lengths.max() if not lengths.empty else 0
                        max_len = max(val_len, len(str(col))) + 2
                        col_letter = get_column_letter(i + 1)
                        worksheet.column_dimensions[col_letter].width = max_len
            
            # Consolidate all validation sheets into one single 'All_Validation' sheet
            all_validation_dfs = []
            
            for (val_generic, val_partname), (val_df, top_3_models) in validation_sheets.items():
                # Filter to only actual_yield, group, and the top 3 models' predictions
                keep_cols = ['actual_yield', 'group']
                for m_name in top_3_models:
                    pred_col = f"{m_name}"  # ML_Train_Model.py saves as just the model name
                    if pred_col not in keep_cols:
                        keep_cols.append(pred_col)
                        
                # Ensure only pick columns that actually exist in the df
                existing_keep_cols = [c for c in keep_cols if c in val_df.columns]
                val_df_filtered = val_df[existing_keep_cols].copy()
                
                # Add columns to identify the source of the validation rows
                val_df_filtered.insert(0, 'Partname', val_partname)
                val_df_filtered.insert(0, 'Generic', val_generic)
                
                all_validation_dfs.append(val_df_filtered)
                
            if all_validation_dfs:
                combined_val_df = pd.concat(all_validation_dfs, ignore_index=True)
                
                # Check for existing 'All_Validation' or 'All_Validations' sheet
                target_sheet_name = 'All_Validation'
                if OUTPUT_FILE.exists() and 'existing_excel_file' in locals():
                    try:
                        # Find existing sheet name (Support both singular/plural)
                        found_sheet = next((s for s in existing_excel_file.sheet_names if s.lower() in ['all_validation', 'all_validations']), None)
                        
                        if found_sheet:
                            target_sheet_name = found_sheet
                            existing_val_df = existing_excel_file.parse(found_sheet)
                            
                            # Filter out old rows for (Generic, Partname) combos are currently processing
                            if 'Generic' in existing_val_df.columns and 'Partname' in existing_val_df.columns:
                                # Create a temporary key for filtering
                                current_combos = combined_val_df[['Generic', 'Partname']].drop_duplicates()
                                
                                merge_temp = pd.merge(existing_val_df, current_combos, on=['Generic', 'Partname'], how='left', indicator=True)
                                existing_val_df_filtered = existing_val_df[merge_temp['_merge'] == 'left_only']
                                
                                combined_val_df = pd.concat([existing_val_df_filtered, combined_val_df], ignore_index=True)
                            elif 'Generic' in existing_val_df.columns:
                                # Fallback if Partname column is missing in old file
                                processed_generic_names = combined_val_df['Generic'].unique()
                                existing_val_df_filtered = existing_val_df[~existing_val_df['Generic'].isin(processed_generic_names)]
                                combined_val_df = pd.concat([existing_val_df_filtered, combined_val_df], ignore_index=True)
                    except Exception as e:
                        print(f"Failed to merge existing validation sheet: {e}")
                
                # Setup styling function for the combined validation sheet
                xl_colors = ["#90EE90", "#ADD8E6", "#FFD580"] # L-Green, L-Blue, L-Orange
                
                def style_excel_combined(st_df):
                    color_df = pd.DataFrame('', index=st_df.index, columns=st_df.columns)
                    meta_cols = {'Generic', 'Partname', 'actual_yield', 'group'}
                    for i in range(len(st_df)):
                        row = st_df.iloc[i]
                        pred_cols_in_row = [c for c in st_df.columns if c not in meta_cols and pd.notna(row[c])]
                        for j, col_name in enumerate(pred_cols_in_row[:3]):
                             if j < len(xl_colors):
                                color_df.iloc[i, st_df.columns.get_loc(col_name)] = f'background-color: {xl_colors[j]}'
                    return color_df
                    
                styled_df = combined_val_df.style.apply(style_excel_combined, axis=None)
                styled_df.to_excel(writer, sheet_name=target_sheet_name, index=False)
                
                # Setup auto-column mapping
                worksheet = writer.sheets[target_sheet_name]
                for i, col in enumerate(combined_val_df.columns):
                    lengths = combined_val_df[col].dropna().astype(str).map(len)
                    val_len = lengths.max() if not lengths.empty else 0
                    max_len = max(val_len, len(str(col))) + 2
                    if i < 40:
                        col_letter = get_column_letter(i + 1)
                        worksheet.column_dimensions[col_letter].width = max_len
                
            # Retain other existing sheets (like user-added ones), but NOT individual generic sheets
            # As per user feedback: "no need, since the details is in all_validation"
            if OUTPUT_FILE.exists() and 'existing_excel_file' in locals():
                try:
                    existing_sheets = existing_excel_file.sheet_names
                    
                    core_sheets = ['Status', 'Summary', 'All_Validation', 'All_Validations']
                    for old_sheet in existing_sheets:
                        if old_sheet not in core_sheets and old_sheet not in writer.sheets: 
                            pass 
                except Exception as e:
                    print(f"Failed to process old sheets: {e}")
                except Exception as e:
                    print(f"Failed to copy old validation sheets: {e}")
                    
        print(f"\nSuccessfully generated {OUTPUT_FILE.name}")
    else:
        print("\nNo devices processed.")

def main():
    extract_model_accuracies()

if __name__ == "__main__":
    main()
